#!C:\Python34
import openpyxl
#import vcf

def GetVCF(name, email, number, org):
	vcf = "BEGIN:VCARD"+"\n"
	vcf += "VERSION:2.1"+"\n"
	vcf += "FN:"+name+"\n"
	vcf += "ORG:"+org+"\n"
	vcf += "TEL;WORK;VOICE:"+str(number)+"\n"
	vcf += "EMAIL:"+email+"\n"
	vcf += "END:VCARD"
	return vcf

def main():
	path = "C:\\1.Class\\PracticeInClass\\Class20\\names_input.xlsx"
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active 
	cell_obj = sheet_obj.cell(row = 1, column = 1)
	#print(cell_obj.value) 
	#print(sheet_obj.max_row) 
	
	max_row=sheet_obj.max_row
	
	max_column=sheet_obj.max_column
	
	#vcf_writer = vcf.Writer(open('C:\\1.Class\\PracticeInClass\\Class20\\contact.vcf', 'w'), vcf_reader)
	fd = open('C:\\1.Class\\PracticeInClass\\Class20\\contact.vcf', 'w')
	
	# iterate over all cells 
	# iterate over all rows
	for i in range(1,max_row+1):
     
     # iterate over all columns
		for j in range(1,max_column+1):
          # get particular cell value    
			cell_obj=sheet_obj.cell(row=i,column=j)
          # print cell value     
			print(cell_obj.value,end=' | ')
			#print(cell_obj.value[0])
			#vcf_writer.write_record(record)
			fd.write("BEGIN:VCARD VERSION:2.1 \n")
			#fd.write(str("FN:" + str(cell_obj.value[0])))
		print('\n')
	fd.close()
	
	

if __name__=="__main__":
	main()